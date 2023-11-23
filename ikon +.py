import tkinter as tk
from tkinter import filedialog as fd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os

class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.workbook = Workbook()
        self.active_sheet = None
        self.init_ui()

    def init_ui(self):
        self.title("Lembar Kerja")
        self.geometry("300x200")

        add_sheet_button = tk.Button(self, text="+", command=self.add_sheet)
        add_sheet_button.pack()

    def add_sheet(self):
        self.active_sheet = self.workbook.active
        self.active_sheet.title = "Sheet"
        for column in range(1, 5):
            self.active_sheet.cell(row=1, column=column).value = ['Tanggal', 'Waktu', 'Kategori', 'Jumlah'][column-1]
            self.active_sheet.cell(row=1, column=column).alignment = Alignment(horizontal='center', vertical='center')

        file_name = "lembar_kerja.xlsx"
        if os.path.exists(file_name):
            self.workbook = Workbook(file_name)
        else:
            self.workbook.save(file_name)

        save_button = tk.Button(self, text="Simpan", command=self.save_sheet)
        save_button.pack()

    def save_sheet(self):
        file_name = fd.asksaveasfilename(defaultextension=".xlsx")
        self.workbook.save(file_name)

app = Application()
app.mainloop()